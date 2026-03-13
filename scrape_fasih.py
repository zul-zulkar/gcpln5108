import asyncio
import openpyxl
from playwright.async_api import async_playwright
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
import requests

# ─── Config ───────────────────────────────────────────────────────────────────
INPUT_FILE = "input/daftar_petugas.xlsx"
OUTPUT_DIR = "output"

# Survey list URL (Angular redirects here after login)
SURVEY_LIST_URL = "https://fasih-sm.bps.go.id/survey-collection/survey"

# Key: output label, Value: direct collect URL
SURVEYS = {
    "pascabayar": "https://fasih-sm.bps.go.id/survey-collection/collect/2e31188c-a617-4163-8056-edccf93d8d79",
    "prabayar":   "https://fasih-sm.bps.go.id/survey-collection/collect/2395b67d-d1af-4739-9ef8-c0cc0aa9ce9a",
}

USERNAME = ""   # diisi saat runtime via GUI/CLI
PASSWORD = ""   # diisi saat runtime via GUI/CLI

UPI_TEXT = "[55]"       # [55] BALI
UP3_TEXT = "[55UTR]"    # [55UTR] BALI UTARA

# Google Apps Script Web App URL (deploy dulu, lalu isi di sini)
SHEETS_WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbz7m9dQ8sl9wqcnqW1P395u7Df1eQ1abZg1VPsZjhe4CBpk6X9_oOjUlUmNfLvkSQiA/exec"

# ──────────────────────────────────────────────────────────────────────────────


def read_petugas():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    petugas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nama, email = row[0], row[1]
        if nama and email:
            petugas.append({"nama": str(nama).strip(), "email": str(email).strip()})
    return petugas


async def login(page):
    """Login via SSO BPS."""
    print(f"[LOGIN] {page.url}")

    try:
        btn = await page.wait_for_selector("a.login-button", timeout=10000)
        await btn.click()
        await page.wait_for_load_state("domcontentloaded", timeout=20000)
        print(f"[LOGIN] SSO redirect: {page.url}")
    except Exception as e:
        print(f"[LOGIN] SSO button error: {e}")

    try:
        user_el = await page.wait_for_selector("input[name='username']", timeout=10000)
        await user_el.fill(USERNAME)
        await page.fill("input[name='password']", PASSWORD)
        await page.click("input[type='submit']")
        await page.wait_for_load_state("domcontentloaded", timeout=20000)
        print(f"[LOGIN] After submit: {page.url}")
    except Exception as e:
        print(f"[LOGIN] Form error: {e}")


async def ngx_select(page, selector, search_text, timeout=10000):
    """Click ngx-select dropdown and choose option containing search_text.
    Handles both static dropdown and type-to-search (server-side) dropdown.
    """
    toggle_sel = f"{selector} .ngx-select__toggle"
    choices_sel = f"{selector} .ngx-select__choices"
    search_input_sel = f"{selector} input.ngx-select__search"

    # Click toggle to open dropdown (dispatchEvent with bubble triggers Angular zone)
    toggle = await page.wait_for_selector(toggle_sel, timeout=timeout)
    await page.evaluate(
        "el => el.dispatchEvent(new MouseEvent('click', {bubbles: true, cancelable: true}))",
        toggle
    )

    # Wait briefly for dropdown to open
    await page.wait_for_timeout(400)

    # If search input exists → type-to-search dropdown (server-side); always type to get correct results
    search_input = await page.query_selector(search_input_sel)
    if search_input:
        await page.evaluate(
            """(el) => {
                el.focus();
                el.value = '';
                el.dispatchEvent(new Event('input', {bubbles: true}));
            }""",
            search_input
        )
        await page.wait_for_timeout(200)
        await page.evaluate(
            """([el, val]) => {
                el.value = val;
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true}));
            }""",
            [search_input, search_text]
        )
        # Wait for server results
        await page.wait_for_timeout(800)
    else:
        # Static dropdown: check items appeared
        item_count = await page.evaluate(
            f"() => document.querySelectorAll('{choices_sel} .ngx-select__item').length"
        )
        if item_count == 0:
            print(f"    [WARN] No items and no search input for {selector}")
            return None

    # Find matching item and click via JS (single evaluate call, no viewport issue)
    clicked = await page.evaluate(
        """([sel, search]) => {
            const items = document.querySelectorAll(sel + ' .ngx-select__item');
            const q = search.toLowerCase();
            for (const item of items) {
                const t = (item.innerText || '').trim();
                if (t.toLowerCase().includes(q)) { item.click(); return t; }
            }
            return null;
        }""",
        [selector, search_text]
    )

    if clicked:
        print(f"    Selected: {clicked}")
        return clicked

    print(f"    [WARN] No match for '{search_text}' in {selector}")
    await page.keyboard.press("Escape")
    return None


async def open_filter(page):
    """Open filter sidebar."""
    btn = await page.wait_for_selector("button:has-text('Filter')", timeout=10000)
    # JS click bypasses viewport restrictions (Angular SPA layout quirks)
    await page.evaluate("el => el.click()", btn)
    try:
        await page.wait_for_selector(".sidebar-content", timeout=4000)
        return True
    except:
        return False


async def close_filter(page):
    """Close filter sidebar by clicking X (JS click — sidebar may be outside viewport)."""
    try:
        close_btn = await page.query_selector(".sidebar-close, button.sidebar-close, [class*='sidebar-close'] button")
        if close_btn:
            await page.evaluate("el => el.click()", close_btn)
            return
        await page.keyboard.press("Escape")
    except:
        pass


async def click_filter_data(page):
    """Click the Filter Data button inside the sidebar and wait for data to reload."""
    btn = await page.wait_for_selector(
        ".sidebar-content button:has-text('Filter Data')", timeout=8000
    )
    await page.evaluate("el => el.click()", btn)
    # Wait for network to settle (Angular XHR), cap at 5s
    try:
        await page.wait_for_load_state("networkidle", timeout=5000)
    except:
        await page.wait_for_timeout(800)


async def extract_stats(page):
    """Extract Open, Submitted by Pencacah, and Rejected by Admin Kabupaten values."""
    open_val = 0
    submitted_val = 0
    rejected_val = 0

    try:
        result = await page.evaluate(r"""() => {
            const stats = {open: 0, submitted: 0, rejected: 0};
            const buttons = document.querySelectorAll('button.btn-outline-primary, button.btn-sm');
            for (const btn of buttons) {
                const text = (btn.innerText || '').trim().toLowerCase();
                const badge = btn.querySelector('span.badge, span.badge-primary, .badge');
                if (!badge) continue;
                const numText = (badge.innerText || badge.textContent || '').trim().replace(/,/g, '');
                const num = parseInt(numText) || 0;
                if (text.includes('open') && !text.includes('submit')) {
                    stats.open = num;
                } else if (text.includes('submit') && text.includes('pencacah')) {
                    stats.submitted = num;
                } else if (text.includes('reject')) {
                    stats.rejected = num;
                }
            }
            return stats;
        }""")

        open_val = result.get("open", 0)
        submitted_val = result.get("submitted", 0)
        rejected_val = result.get("rejected", 0)

    except Exception as e:
        print(f"    [ERROR] extract_stats: {e}")

    return open_val, submitted_val, rejected_val


async def scrape_survey(page, survey_name, url, petugas_list, stop_event=None):
    """Scrape one survey for all petugas."""
    print(f"\n[SURVEY] {survey_name.upper()}")

    # Navigasi langsung ke URL survey (lebih reliabel di headless mode)
    collect_id = url.split("/")[-1]
    if collect_id not in page.url:
        print(f"[NAV] Navigating to: {url}")
        await page.goto(url, wait_until="domcontentloaded", timeout=40000)
        # Jika di-redirect ke login (session expired), login ulang
        if "fasih-sm.bps.go.id" not in page.url or "oauth" in page.url or "sso" in page.url:
            print(f"[LOGIN] Session expired, re-logging in... ({page.url})")
            await login(page)
            # Setelah login, navigate ulang ke survey
            print("[NAV] Re-navigating after login...")
            await page.goto(url, wait_until="domcontentloaded", timeout=40000)
        try:
            await page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            await page.wait_for_timeout(2000)
        await page.wait_for_timeout(800)

    print(f"[NAV] {page.url}")

    # Click "Data" tab to ensure filter is visible (JS click — may be outside viewport)
    try:
        data_tab = await page.wait_for_selector("li.nav-item a:has-text('Data')", timeout=10000)
        await page.evaluate("el => el.click()", data_tab)
        await page.wait_for_timeout(500)
    except:
        pass  # might already be on Data tab

    try:
        await page.wait_for_selector("button:has-text('Filter')", timeout=25000)
        print("[INFO] Page loaded OK")
    except:
        print("[WARN] Filter button not found")
        await page.screenshot(path=f"output/debug_{survey_name}.png")
        print(f"[DEBUG] Screenshot saved: output/debug_{survey_name}.png")

    print("\n" + "=" * 75)
    print(f"{'No':>4}  {'Nama':<35} {'Open':>6} {'Submitted':>10} {'Rejected':>9}")
    print("=" * 75)

    results = []

    # Open filter once, set UPI/UP3 once — keep sidebar open for all records
    opened = await open_filter(page)
    if not opened:
        print("[ERROR] Could not open filter sidebar")
        return []

    await ngx_select(page, 'ngx-select[name="region1Id"]', UPI_TEXT)
    await ngx_select(page, 'ngx-select[name="region2Id"]', UP3_TEXT)

    # Click Filter Data once with UPI/UP3 to trigger initial data load,
    # letting Angular fully initialize before starting the petugas loop
    print("[INIT] Applying UPI/UP3 filter to pre-load data...")
    await click_filter_data(page)
    # Extra settle time after initial load
    try:
        await page.wait_for_load_state("networkidle", timeout=5000)
    except:
        await page.wait_for_timeout(1500)
    print("[INIT] Initial data load done, starting petugas loop")

    for idx, pt in enumerate(petugas_list):
        if stop_event and stop_event.is_set():
            print("\n[STOP] Dihentikan oleh pengguna.")
            break

        nama = pt["nama"]
        email = pt["email"]

        try:
            # Ensure sidebar is still open
            sidebar_visible = await page.evaluate(
                "() => { const s = document.querySelector('.sidebar-content'); return s ? s.offsetParent !== null || s.getBoundingClientRect().width > 0 : false; }"
            )
            if not sidebar_visible:
                opened = await open_filter(page)
                if not opened:
                    print(f"[{idx+1:02d}] {nama[:30]:<30} - FILTER_ERR")
                    results.append({"nama": nama, "email": email, "open": "ERR", "submitted": "ERR", "rejected": "ERR"})
                    continue

            # Retry up to 3x if pencacah not matched
            selected = None
            for attempt in range(3):
                selected = await ngx_select(page, 'ngx-select[optionvaluefield="username"]', email)
                if selected:
                    break
                print(f"    [RETRY {attempt+1}/3] {email}")
                await page.wait_for_timeout(500)

            if not selected:
                print(f"[{idx+1:02d}] {nama[:35]:<35} SKIP (no match after 3 attempts)")
                results.append({"nama": nama, "email": email, "open": "NO_MATCH", "submitted": "NO_MATCH", "rejected": "NO_MATCH"})
                continue

            # Wait for Angular to settle after selection (may trigger async HTTP on type-to-search)
            try:
                await page.wait_for_load_state("networkidle", timeout=2000)
            except:
                await page.wait_for_timeout(500)

            # Verify toggle no longer shows placeholder before applying filter
            for _ in range(15):
                toggle_text = await page.evaluate(
                    "() => { const t = document.querySelector('ngx-select[optionvaluefield=\"username\"] .ngx-select__toggle'); return t ? (t.innerText || '').trim() : ''; }"
                )
                if toggle_text and toggle_text.lower() not in ("select...", "select", ""):
                    break
                await page.wait_for_timeout(200)

            # Snapshot stats BEFORE clicking filter — used to detect DOM update
            stats_before = await extract_stats(page)

            await click_filter_data(page)

            # Verify filter was applied correctly: toggle must still show selected pencacah.
            # If it reverted to placeholder, the filter ran without pencacah — re-select and retry.
            for verify_attempt in range(3):
                post_toggle = await page.evaluate(
                    "() => { const t = document.querySelector('ngx-select[optionvaluefield=\"username\"] .ngx-select__toggle'); return t ? (t.innerText || '').trim() : ''; }"
                )
                if post_toggle and post_toggle.lower() not in ("select...", "select", ""):
                    break
                print(f"    [FILTER-VERIFY] Toggle reverted to placeholder (attempt {verify_attempt+1}/3), re-selecting...")
                await page.wait_for_timeout(500)
                re_sel = await ngx_select(page, 'ngx-select[optionvaluefield="username"]', email)
                if re_sel:
                    try:
                        await page.wait_for_load_state("networkidle", timeout=2000)
                    except:
                        await page.wait_for_timeout(500)
                    await click_filter_data(page)
                else:
                    await page.wait_for_timeout(500)

            # Poll until stats badge values actually change in the DOM (max 4s).
            # This guarantees Angular has re-rendered the new petugas data,
            # not still showing the previous filter's values (root cause of [01] off-by-one).
            open_val, submitted_val, rejected_val = stats_before
            for _ in range(20):
                open_val, submitted_val, rejected_val = await extract_stats(page)
                if (open_val, submitted_val, rejected_val) != stats_before:
                    break
                await page.wait_for_timeout(200)
            results.append({"nama": nama, "email": email, "open": open_val, "submitted": submitted_val, "rejected": rejected_val})
            print(f"[{idx+1:02d}] {nama[:35]:<35} {open_val:>6} {submitted_val:>10} {rejected_val:>9}")

        except Exception as e:
            print(f"[{idx+1:02d}] {nama[:35]:<35} ERROR: {e}")
            results.append({"nama": nama, "email": email, "open": "ERR", "submitted": "ERR", "rejected": "ERR"})
            try:
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(300)
                await open_filter(page)
            except:
                pass

        if (idx + 1) % 50 == 0:
            save_results(results, survey_name)

    save_results(results, survey_name)
    await close_filter(page)
    print("=" * 75)
    print(f"[DONE] {survey_name}: {len(results)} petugas")
    return results


def save_results(results, label="rekap"):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUTPUT_DIR, f"rekap_fasih_{label}_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap"

    headers = ["No", "Nama", "Email", "Open", "Submitted by Pencacah", "Rejected by Admin Kabupaten"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(results, 1):
        ws.append([i, r["nama"], r["email"], r["open"], r["submitted"], r.get("rejected", 0)])
        if i % 2 == 0:
            for cell in ws[i + 1]:
                cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    col_widths = {"A": 6, "B": 35, "C": 38, "D": 10, "E": 25, "F": 30}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22

    wb.save(out_path)
    print(f"  [SAVED] {out_path}")
    return out_path


def append_detail_snapshot(all_results):
    """Kirim data per pencacah ke Google Sheets (tab Riwayat) via Apps Script webhook."""
    if not SHEETS_WEBHOOK_URL:
        return

    pasca = {r["email"]: r for r in all_results.get("pascabayar", [])}
    praba = {r["email"]: r for r in all_results.get("prabayar", [])}
    all_emails = sorted(set(list(pasca.keys()) + list(praba.keys())))

    now  = datetime.now()
    rows = []
    for email in all_emails:
        rp = pasca.get(email, {})
        rr = praba.get(email, {})
        rows.append({
            "nama":         rp.get("nama") or rr.get("nama", ""),
            "email":        email,
            "open_pasca":   rp.get("open"),
            "submit_pasca": rp.get("submitted"),
            "reject_pasca": rp.get("rejected"),
            "open_praba":   rr.get("open"),
            "submit_praba": rr.get("submitted"),
            "reject_praba": rr.get("rejected"),
        })

    payload = {
        "type":    "detail",
        "tanggal": now.strftime("%d/%m/%Y"),
        "waktu":   now.strftime("%H:%M"),
        "rows":    rows,
    }

    try:
        resp = requests.post(SHEETS_WEBHOOK_URL, json=payload, timeout=30)
        print(f"[DETAIL] HTTP {resp.status_code} → {resp.text[:300]}")
    except Exception as e:
        print(f"[DETAIL] Gagal kirim: {e}")


def append_daily_snapshot(all_results):
    """Kirim ringkasan harian ke Google Sheets via Apps Script webhook."""
    if not SHEETS_WEBHOOK_URL:
        print("[SNAPSHOT] SHEETS_WEBHOOK_URL belum diset, snapshot dilewati.")
        return

    def sum_num(results, key):
        return sum(r[key] for r in results if isinstance(r.get(key), (int, float)))

    pasca = all_results.get("pascabayar", [])
    praba = all_results.get("prabayar", [])
    now   = datetime.now()

    payload = {
        "tanggal":      now.strftime("%d/%m/%Y"),
        "waktu":        now.strftime("%H:%M"),
        "open_pasca":   sum_num(pasca, "open"),
        "submit_pasca": sum_num(pasca, "submitted"),
        "reject_pasca": sum_num(pasca, "rejected"),
        "open_praba":   sum_num(praba, "open"),
        "submit_praba": sum_num(praba, "submitted"),
        "reject_praba": sum_num(praba, "rejected"),
    }

    try:
        resp = requests.post(SHEETS_WEBHOOK_URL, json=payload, timeout=15)
        print(
            f"[SNAPSHOT] HTTP {resp.status_code} → {resp.text[:300]}\n"
            f"  {payload['tanggal']} {payload['waktu']}  "
            f"Pasca submit={payload['submit_pasca']}  Praba submit={payload['submit_praba']}"
        )
    except Exception as e:
        print(f"[SNAPSHOT] Gagal kirim: {e}")


async def main():
    petugas_list = read_petugas()
    print(f"[INFO] {len(petugas_list)} petugas loaded")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, slow_mo=50)
        context = await browser.new_context(viewport={"width": 1400, "height": 900})
        page = await context.new_page()

        # Login once via survey list URL (SSO redirects here)
        print(f"[INFO] Opening: {SURVEY_LIST_URL}")
        await page.goto(SURVEY_LIST_URL, wait_until="domcontentloaded", timeout=30000)

        if SURVEY_LIST_URL not in page.url:
            await login(page)

        print(f"[INFO] Logged in: {page.url}")

        all_results = {}
        for survey_name, url in SURVEYS.items():
            results = await scrape_survey(page, survey_name, url, petugas_list)
            all_results[survey_name] = results

        await browser.close()

    append_daily_snapshot(all_results)
    append_detail_snapshot(all_results)


async def main_with_stop(stop_event, input_file=None, username=None, password=None, headless=False, sheets_url=None):
    """Entry point untuk GUI — menerima parameter dinamis dan stop_event."""
    global INPUT_FILE, USERNAME, PASSWORD, SHEETS_WEBHOOK_URL
    if input_file:
        INPUT_FILE = input_file
    if username:
        USERNAME = username
    if password:
        PASSWORD = password
    if sheets_url is not None:
        SHEETS_WEBHOOK_URL = sheets_url

    petugas_list = read_petugas()
    print(f"[INFO] {len(petugas_list)} petugas loaded dari {INPUT_FILE}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless, slow_mo=50)
        context = await browser.new_context(viewport={"width": 1400, "height": 900})
        page = await context.new_page()

        print(f"[INFO] Opening: {SURVEY_LIST_URL}")
        await page.goto(SURVEY_LIST_URL, wait_until="domcontentloaded", timeout=30000)

        if SURVEY_LIST_URL not in page.url:
            await login(page)
            # Tunggu Angular survey list selesai render setelah redirect SSO
            print("[INFO] Waiting for survey list to load...")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                await page.wait_for_timeout(4000)

        print(f"[INFO] Logged in: {page.url}")

        all_results = {}
        for survey_name, url in SURVEYS.items():
            if stop_event and stop_event.is_set():
                break
            results = await scrape_survey(page, survey_name, url, petugas_list, stop_event)
            all_results[survey_name] = results

        await browser.close()
        print("[INFO] Browser ditutup.")

    if not (stop_event and stop_event.is_set()):
        append_daily_snapshot(all_results)
        append_detail_snapshot(all_results)


if __name__ == "__main__":
    asyncio.run(main())
